"""
Report Generation Module
Generates PDF and Excel reports with analytics data
"""

from datetime import datetime, timedelta
from io import BytesIO
import json
from pathlib import Path

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReportGenerator:
    """Generates reports in multiple formats (PDF, Excel, JSON)"""
    
    def __init__(self):
        self.report_data = {}
        
    def set_report_data(self, period="weekly", metrics=None, trends=None, 
                       top_performers=None, insights=None, alerts=None):
        """Set the data for report generation"""
        self.report_data = {
            "period": period,
            "generated_at": datetime.now().isoformat(),
            "metrics": metrics or self._get_default_metrics(),
            "trends": trends or self._get_default_trends(),
            "top_performers": top_performers or self._get_default_performers(),
            "insights": insights or self._get_default_insights(),
            "alerts": alerts or self._get_default_alerts(),
        }
        
    def _get_default_metrics(self):
        """Get default metrics"""
        return {
            "total_visits": 18543,
            "avg_attendance": 87.5,
            "total_engineers": 20,
            "offline_sites": 2134,
            "completion_rate": 89.2,
            "avg_tickets_per_visit": 2.34,
        }
        
    def _get_default_trends(self):
        """Get default trends"""
        trends = []
        for i in range(30, 0, -1):
            date = datetime.now() - timedelta(days=i)
            trends.append({
                "date": date.strftime("%Y-%m-%d"),
                "visits": 18543 // 30 * i,
                "attendance": 87.5 + (i % 5),
            })
        return trends
        
    def _get_default_performers(self):
        """Get default top performers"""
        return [
            {"rank": 1, "name": "Rajesh Kumar", "code": "001", "visits": 145, "attendance": 94.5, "completion": 92.3},
            {"rank": 2, "name": "Priya Singh", "code": "002", "visits": 138, "attendance": 96.2, "completion": 95.1},
            {"rank": 3, "name": "Amit Patel", "code": "003", "visits": 125, "attendance": 88.9, "completion": 87.6},
            {"rank": 4, "name": "Neha Verma", "code": "004", "visits": 118, "attendance": 91.2, "completion": 90.4},
            {"rank": 5, "name": "Vikram Singh", "code": "005", "visits": 112, "attendance": 85.3, "completion": 84.2},
        ]
        
    def _get_default_insights(self):
        """Get default insights"""
        return [
            {
                "type": "alert",
                "severity": "high",
                "message": "45 sites offline for >60 days. Urgent attention needed.",
                "metric": "offline_sites",
            },
            {
                "type": "positive",
                "severity": "low",
                "message": "Attendance improved by 3.2% this week vs last week.",
                "metric": "attendance",
            },
            {
                "type": "warning",
                "severity": "medium",
                "message": "3 engineers below 70% ticket closure rate.",
                "metric": "closure_rate",
            },
            {
                "type": "positive",
                "severity": "low",
                "message": "Visit volume up 12% vs last month.",
                "metric": "visits",
            },
        ]
        
    def _get_default_alerts(self):
        """Get default alerts"""
        return [
            {"severity": "high", "message": "45 sites offline for >60 days"},
            {"severity": "medium", "message": "3 engineers below 70% completion rate"},
            {"severity": "low", "message": "5 sites with irregular visit patterns"},
        ]

    def generate_json(self):
        """Generate JSON report"""
        return json.dumps(self.report_data, indent=2, default=str)

    def generate_excel(self):
        """Generate Excel workbook"""
        if not OPENPYXL_AVAILABLE:
            raise Exception("openpyxl not installed. Install with: pip install openpyxl")
            
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Summary Sheet
        ws_summary = wb.create_sheet("Summary", 0)
        self._create_summary_sheet(ws_summary)
        
        # Metrics Sheet
        ws_metrics = wb.create_sheet("Metrics", 1)
        self._create_metrics_sheet(ws_metrics)
        
        # Performers Sheet
        ws_performers = wb.create_sheet("Top Performers", 2)
        self._create_performers_sheet(ws_performers)
        
        # Insights Sheet
        ws_insights = wb.create_sheet("Insights", 3)
        self._create_insights_sheet(ws_insights)
        
        # Trends Sheet
        ws_trends = wb.create_sheet("Trends", 4)
        self._create_trends_sheet(ws_trends)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _create_summary_sheet(self, ws):
        """Create summary sheet"""
        # Header
        ws['A1'] = "VProtect Field Service Dashboard - Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Period: {self.report_data['period'].capitalize()}"
        ws['A2'].font = Font(size=12)
        ws['A3'] = f"Generated: {self.report_data['generated_at']}"
        
        # Metrics summary
        row = 5
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        
        metrics = self.report_data['metrics']
        row = 6
        for key, value in metrics.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
            
        # Column width
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _create_metrics_sheet(self, ws):
        """Create detailed metrics sheet"""
        metrics = self.report_data['metrics']
        
        # Header
        ws['A1'] = "Detailed Metrics"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Data
        row = 3
        for key, value in metrics.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = value
            row += 1
            
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _create_performers_sheet(self, ws):
        """Create top performers sheet"""
        # Header
        ws['A1'] = "Top 5 Performing Engineers"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        
        # Column headers
        headers = ["Rank", "Engineer Name", "Code", "Total Visits", "Attendance %", "Completion Rate %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
        # Data
        for row_idx, performer in enumerate(self.report_data['top_performers'], 4):
            ws.cell(row=row_idx, column=1).value = performer['rank']
            ws.cell(row=row_idx, column=2).value = performer['name']
            ws.cell(row=row_idx, column=3).value = performer['code']
            ws.cell(row=row_idx, column=4).value = performer['visits']
            ws.cell(row=row_idx, column=5).value = performer['attendance']
            ws.cell(row=row_idx, column=6).value = performer['completion']
            
        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18

    def _create_insights_sheet(self, ws):
        """Create insights sheet"""
        # Header
        ws['A1'] = "Analytics Insights & Alerts"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Column headers
        headers = ["Type", "Severity", "Message", "Metric"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
            
        # Data
        for row_idx, insight in enumerate(self.report_data['insights'], 4):
            ws.cell(row=row_idx, column=1).value = insight.get('type', '')
            ws.cell(row=row_idx, column=2).value = insight.get('severity', '')
            ws.cell(row=row_idx, column=3).value = insight.get('message', '')
            ws.cell(row=row_idx, column=4).value = insight.get('metric', '')
            
            # Color severity
            if insight.get('severity') == 'high':
                ws.cell(row=row_idx, column=2).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws.cell(row=row_idx, column=2).font = Font(color="FFFFFF", bold=True)
            elif insight.get('severity') == 'medium':
                ws.cell(row=row_idx, column=2).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15

    def _create_trends_sheet(self, ws):
        """Create trends sheet"""
        # Header
        ws['A1'] = "Trend Data"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Column headers
        headers = ["Date", "Visits", "Attendance %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            
        # Data
        for row_idx, trend in enumerate(self.report_data['trends'], 4):
            ws.cell(row=row_idx, column=1).value = trend.get('date', '')
            ws.cell(row=row_idx, column=2).value = trend.get('visits', 0)
            ws.cell(row=row_idx, column=3).value = trend.get('attendance', 0)
            
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    def generate_pdf(self):
        """Generate PDF report"""
        if not REPORTLAB_AVAILABLE:
            raise Exception("reportlab not installed. Install with: pip install reportlab")
            
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#0066CC'),
            spaceAfter=30,
            alignment=1,
        )
        story.append(Paragraph("VProtect Dashboard Report", title_style))
        
        # Period info
        period_text = f"Period: {self.report_data['period'].capitalize()} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        story.append(Paragraph(period_text, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Metrics Table
        story.append(Paragraph("Key Metrics", styles['Heading2']))
        metrics_data = [['Metric', 'Value']]
        for key, value in self.report_data['metrics'].items():
            metrics_data.append([key.replace('_', ' ').title(), str(value)])
            
        metrics_table = Table(metrics_data, colWidths=[3*inch, 2*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(metrics_table)
        story.append(Spacer(1, 0.5*inch))
        
        # Top Performers
        story.append(PageBreak())
        story.append(Paragraph("Top 5 Performing Engineers", styles['Heading2']))
        performers_data = [['Rank', 'Name', 'Code', 'Visits', 'Attendance', 'Completion']]
        for perf in self.report_data['top_performers']:
            performers_data.append([
                str(perf['rank']),
                perf['name'],
                perf['code'],
                str(perf['visits']),
                f"{perf['attendance']:.1f}%",
                f"{perf['completion']:.1f}%",
            ])
            
        performers_table = Table(performers_data, colWidths=[0.8*inch, 1.5*inch, 0.7*inch, 1*inch, 1.2*inch, 1.2*inch])
        performers_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(performers_table)
        story.append(Spacer(1, 0.5*inch))
        
        # Insights & Alerts
        story.append(Paragraph("Key Insights", styles['Heading2']))
        for insight in self.report_data['insights']:
            insight_text = f"<b>{insight['message']}</b>"
            story.append(Paragraph(insight_text, styles['Normal']))
            story.append(Spacer(1, 0.1*inch))
            
        # Build PDF
        doc.build(story)
        output.seek(0)
        return output.getvalue()
