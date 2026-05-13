"""
FastAPI Backend — Service Analysis Dashboard
All API routes: auth, upload, dashboard KPIs, engineers, tickets, fraud, states
"""

import os, io, json, shutil, hashlib
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd
from fastapi import FastAPI, UploadFile, File, Depends, HTTPException, Query, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine
from pydantic import BaseModel
import bcrypt
import jwt as pyjwt

# Import ETL pipeline
import sys
sys.path.append(str(Path(__file__).parent.parent))
from etl.pipeline import (
    get_engine, run_full_etl, start_etl_job, get_job_status,
    load_b2b_offline, load_offline_data, load_view_ticket,
    load_attendance, load_attendance_data, load_engineer_master, load_site_master,
    load_visit_master, load_service_area_master,
    validate_b2b_offline, validate_view_ticket, validate_attendance,
    validate_engineer, validate_site_master, validate_visit_master,
    ValidationResult
)

# ─── CONFIG ───────────────────────────────────────────────
SECRET_KEY = os.getenv("JWT_SECRET", "change-this-secret-key-in-production-please")
ALGORITHM  = "HS256"
TOKEN_EXPIRE_HOURS = 8
UPLOAD_DIR = Path("data/uploads")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="Service Analysis Dashboard API", version="2.0")

app.add_middleware(CORSMiddleware,
    allow_origins=["*"], allow_credentials=True,
    allow_methods=["*"], allow_headers=["*"])

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")

# ─── AUTH ─────────────────────────────────────────────────
def create_token(user_id: int, username: str, role: str, state_code: Optional[str]) -> str:
    exp = datetime.utcnow() + timedelta(hours=TOKEN_EXPIRE_HOURS)
    return pyjwt.encode({
        "sub": str(user_id), "username": username,
        "role": role, "state_code": state_code, "exp": exp
    }, SECRET_KEY, algorithm=ALGORITHM)


def verify_token(token: str = Depends(oauth2_scheme)) -> dict:
    try:
        payload = pyjwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return payload
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")


def require_ops_or_admin(user=Depends(verify_token)):
    if user.get("role") not in ("ops_manager", "admin"):
        raise HTTPException(status_code=403, detail="Access denied")
    return user


def state_filter(user: dict) -> Optional[str]:
    """Return state restriction for state managers, None for ops/admin"""
    if user.get("role") == "state_manager":
        return user.get("state_code")
    return None


def get_db() -> Engine:
    return get_engine()


# ─── AUTH ENDPOINTS ───────────────────────────────────────
@app.post("/api/auth/login")
async def login(form: OAuth2PasswordRequestForm = Depends(), db: Engine = Depends(get_db)):
    with db.connect() as conn:
        user = conn.execute(text("""
            SELECT id, username, password_hash, role, state_code, force_password_change, is_active
            FROM app_user WHERE username = :u
        """), {"u": form.username}).fetchone()

    if not user or not user.is_active:
        raise HTTPException(status_code=401, detail="Invalid username or password")

    if not bcrypt.checkpw(form.password.encode(), user.password_hash.encode()):
        raise HTTPException(status_code=401, detail="Invalid username or password")

    with db.begin() as conn:
        conn.execute(text("UPDATE app_user SET last_login = NOW() WHERE id = :id"), {"id": user.id})

    token = create_token(user.id, user.username, user.role, user.state_code)
    return {
        "access_token": token, "token_type": "bearer",
        "role": user.role, "state_code": user.state_code,
        "force_password_change": user.force_password_change,
        "username": user.username
    }


@app.post("/api/auth/change-password")
async def change_password(
    old_password: str = Form(...), new_password: str = Form(...),
    user=Depends(verify_token), db: Engine = Depends(get_db)
):
    if len(new_password) < 10:
        raise HTTPException(400, "Password must be at least 10 characters")

    with db.connect() as conn:
        u = conn.execute(text("SELECT password_hash FROM app_user WHERE username = :u"),
                         {"u": user["username"]}).fetchone()
    if not bcrypt.checkpw(old_password.encode(), u.password_hash.encode()):
        raise HTTPException(400, "Old password incorrect")

    new_hash = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()
    with db.begin() as conn:
        conn.execute(text("""
            UPDATE app_user SET password_hash = :ph, force_password_change = FALSE
            WHERE username = :u
        """), {"ph": new_hash, "u": user["username"]})
    return {"message": "Password changed successfully"}


# ─── DATA UPLOAD ENDPOINTS ────────────────────────────────
@app.post("/api/upload/{file_type}")
async def upload_file(
    file_type: str,
    file: UploadFile = File(...),
    user=Depends(verify_token),
    db: Engine = Depends(get_db)
):
    if user.get("role") not in ("admin",):
        raise HTTPException(403, "Only admins can upload files")

    valid_types = ["b2b_offline", "view_ticket", "attendance", "engineer", "site_master", "visit_master", "service_area_master"]
    if file_type not in valid_types:
        raise HTTPException(400, f"Invalid file type. Must be one of: {valid_types}")

    content = await file.read()
    today = date.today()

    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content), encoding='utf-8', low_memory=False)
        elif file_type == "b2b_offline":
            # Data is in sheet named "B2B"; first row may be blank, real header in row 0
            xl = pd.ExcelFile(io.BytesIO(content), engine='openpyxl')
            sheet = "B2B" if "B2B" in xl.sheet_names else xl.sheet_names[0]
            raw = xl.parse(sheet, header=None)
            # Find the row containing 'cs_no' or 'cs no' to use as header
            header_row = 0
            for i, row in raw.iterrows():
                if any(str(v).lower().strip() in ('cs_no','cs no','cs id','cs_id') for v in row.values if pd.notna(v)):
                    header_row = i
                    break
            df = xl.parse(sheet, header=header_row)
            # Drop fully empty columns
            df = df.dropna(axis=1, how='all')
        else:
            df = pd.read_excel(io.BytesIO(content), engine='openpyxl')
    except Exception as e:
        raise HTTPException(400, f"Could not read file: {e}")

    # Validate
    if file_type == "b2b_offline":
        vres = validate_b2b_offline(df)
    elif file_type == "view_ticket":
        with db.connect() as conn:
            prev = conn.execute(text("SELECT MAX(row_count) FROM upload_log WHERE file_type='VIEW_TICKET' AND validation_status='SUCCESS'")).scalar() or 0
        vres = validate_view_ticket(df, db, int(prev))
    elif file_type == "attendance":
        vres = validate_attendance(df, db)
    elif file_type == "engineer":
        vres = validate_engineer(df)
    elif file_type == "site_master":
        vres = validate_site_master(df)
    elif file_type == "visit_master":
        vres = validate_visit_master(df)
    elif file_type == "service_area_master":
        vres = ValidationResult(); vres.row_count = len(df)

    if not vres.passed:
        return JSONResponse(status_code=422, content={
            "status": "FAILED",
            "errors": vres.errors,
            "warnings": vres.warnings,
            "row_count": vres.row_count
        })

    # Save file
    safe_name = f"{file_type}_{today.isoformat()}_{file.filename}"
    save_path = UPLOAD_DIR / safe_name
    with open(save_path, 'wb') as f:
        f.write(content)

    # Log upload
    with db.begin() as conn:
        log_id = conn.execute(text("""
            INSERT INTO upload_log (file_type, original_filename, uploaded_by, row_count,
                validation_status, validation_warnings, stored_filename, upload_date, etl_status)
            VALUES (:ft, :fn, :ub, :rc, 'SUCCESS', cast(:warns as jsonb), :sf, :ud, 'PENDING')
            RETURNING id
        """), {
            "ft": file_type.upper(), "fn": file.filename, "ub": user["username"],
            "rc": vres.row_count, "warns": json.dumps(vres.warnings),
            "sf": safe_name, "ud": today
        }).scalar()

    # Load data then trigger async ETL
    try:
        if file_type == "b2b_offline":
            # Parse data_date from filename if possible (format: B2B Offline DD-MM-YYYY.xlsx)
            import re as _re
            m = _re.search(r'(\d{2})[-_](\d{2})[-_](\d{4})', file.filename or '')
            if m:
                try:
                    data_date = date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
                except: data_date = today
            else:
                data_date = today
            rows = load_offline_data(df, db, data_date)
        elif file_type == "view_ticket":
            rows = load_view_ticket(df, db, today)
        elif file_type == "attendance":
            rows = load_attendance_data(df, db)
        elif file_type == "engineer":
            rows = load_engineer_master(df, db)
        elif file_type == "site_master":
            rows = load_site_master(df, db)
        elif file_type == "visit_master":
            rows = load_visit_master(df, db)
        elif file_type == "service_area_master":
            rows = load_service_area_master(df, db)

        # Start async ETL — returns job_id immediately
        with db.begin() as conn:
            conn.execute(text("UPDATE upload_log SET etl_status='RUNNING', etl_started_at=NOW() WHERE id=:id"), {"id": log_id})

        job_id = start_etl_job(db, today)

        with db.begin() as conn:
            conn.execute(text("UPDATE upload_log SET job_id=:jid WHERE id=:id"), {"jid": job_id, "id": log_id})

        return {
            "status": "SUCCESS",
            "rows_loaded": rows,
            "warnings": vres.warnings,
            "etl_status": "RUNNING",
            "job_id": job_id,
            "upload_id": log_id
        }

    except Exception as e:
        with db.begin() as conn:
            conn.execute(text("UPDATE upload_log SET etl_status='FAILED', etl_errors=:e WHERE id=:id"),
                         {"e": str(e), "id": log_id})
        raise HTTPException(500, f"Data loading error: {e}")


@app.get("/api/upload/status/{job_id}")
async def etl_job_status(job_id: str, user=Depends(verify_token)):
    """Poll ETL job completion status. Frontend polls this after upload."""
    status = get_job_status(job_id)
    if status.get("status") == "NOT_FOUND":
        raise HTTPException(404, "Job not found")
    return status


@app.get("/api/upload/history")
async def upload_history(user=Depends(verify_token), db: Engine = Depends(get_db)):
    with db.connect() as conn:
        rows = conn.execute(text("""
            SELECT id, file_type, original_filename, upload_datetime, uploaded_by,
                   row_count, validation_status, etl_status, upload_date
            FROM upload_log ORDER BY upload_datetime DESC LIMIT 50
        """)).fetchall()
    return [dict(r._mapping) for r in rows]


# ─── DASHBOARD KPI ENDPOINTS ──────────────────────────────
@app.get("/api/dashboard/kpis")
async def dashboard_kpis(
    fact_date: Optional[str] = None,
    user=Depends(verify_token),
    db: Engine = Depends(get_db)
):
    d = date.fromisoformat(fact_date) if fact_date else date.today()
    state_restrict = state_filter(user)

    with db.connect() as conn:
        state_clause = "AND state = :sr" if state_restrict else ""
        params = {"d": d, "sr": state_restrict}

        kpi = conn.execute(text(f"""
            SELECT
                SUM(offline_sites) as total_offline,
                SUM(total_sites) as total_sites,
                SUM(sites_bucket_31_60 + sites_bucket_61_90 + sites_bucket_90plus) as offline_30plus,
                SUM(open_tickets) as open_tickets,
                SUM(pending_tickets) as pending_tickets,
                SUM(engineers_present) as engineers_present,
                SUM(engineers_total) as engineers_total,
                SUM(no_ticket_offline_sites) as no_ticket_offline,
                ROUND(AVG(offline_rate), 2) as avg_offline_rate
            FROM fact_state_daily
            WHERE fact_date = :d {state_clause}
        """), params).fetchone()

        # MTD closure rate
        ym = d.strftime("%Y-%m")
        params["ym"] = ym
        closure = conn.execute(text(f"""
            SELECT ROUND(AVG(closure_rate), 2) as avg_closure
            FROM fact_engineer_monthly
            WHERE year_month = :ym
        """), params).fetchone()

        # Avg aging - open tickets
        avg_aging = conn.execute(text(f"""
            SELECT ROUND(AVG(aging_days), 1) as avg_aging
            FROM view_ticket
            WHERE upload_date = :d AND ticket_status = 'OPEN'
            {'AND site_state_master = :sr' if state_restrict else ''}
        """), params).scalar()

        # Active fraud flags this month
        month_start = d.replace(day=1)
        fraud_count = conn.execute(text(f"""
            SELECT COUNT(*) FROM fraud_flag
            WHERE flag_date >= :ms AND review_status = 'PENDING'
            {'AND state = :sr' if state_restrict else ''}
        """), {**params, "ms": month_start}).scalar()

    return {
        "date": str(d),
        "total_offline_sites": int(kpi.total_offline or 0),
        "total_sites": int(kpi.total_sites or 0),
        "offline_rate_pct": float(kpi.avg_offline_rate or 0),
        "offline_30plus": int(kpi.offline_30plus or 0),
        "open_tickets": int(kpi.open_tickets or 0),
        "pending_tickets": int(kpi.pending_tickets or 0),
        "engineers_present": int(kpi.engineers_present or 0),
        "engineers_total": int(kpi.engineers_total or 0),
        "attendance_rate_pct": round(int(kpi.engineers_present or 0) / max(int(kpi.engineers_total or 1), 1) * 100, 1),
        "mtd_closure_rate": float(closure.avg_closure or 0),
        "avg_open_aging_days": float(avg_aging or 0),
        "active_fraud_flags": int(fraud_count or 0),
        "no_ticket_offline_sites": int(kpi.no_ticket_offline or 0),
    }


@app.get("/api/dashboard/smart-insights")
async def smart_insights(
    fact_date: Optional[str] = None,
    user=Depends(verify_token),
    db: Engine = Depends(get_db)
):
    d = date.fromisoformat(fact_date) if fact_date else date.today()
    state_restrict = state_filter(user)

    with db.connect() as conn:
        state_clause = "AND (state = :sr OR state IS NULL)" if state_restrict else ""
        params = {"d": d, "sr": state_restrict}
        rows = conn.execute(text(f"""
            SELECT rule_id, priority, insight_text, data_payload, state
            FROM smart_insight
            WHERE insight_date = :d AND is_active = TRUE {state_clause}
            ORDER BY CASE priority WHEN 'CRITICAL' THEN 1 WHEN 'WARNING' THEN 2 ELSE 3 END
        """), params).fetchall()
    return [dict(r._mapping) for r in rows]


@app.get("/api/dashboard/state-health")
async def state_health(
    fact_date: Optional[str] = None,
    user=Depends(verify_token),
    db: Engine = Depends(get_db)
):
    d = date.fromisoformat(fact_date) if fact_date else date.today()
    state_restrict = state_filter(user)

    with db.connect() as conn:
        params = {"d": d, "sr": state_restrict}
        state_clause = "AND fsd.state = :sr" if state_restrict else ""
        rows = conn.execute(text(f"""
            SELECT fsd.state, fsd.zone, fsd.offline_rate, fsd.total_sites,
                   fsd.offline_sites, fsd.open_tickets, fsd.engineers_present,
                   fsd.engineers_total, fsd.no_ticket_offline_sites,
                   fsd.sites_bucket_1_3, fsd.sites_bucket_4_7, fsd.sites_bucket_8_15,
                   fsd.sites_bucket_16_30, fsd.sites_bucket_31_60, fsd.sites_bucket_61_90,
                   fsd.sites_bucket_90plus,
                   fsm.state_score, fsm.avg_engineer_score
            FROM fact_state_daily fsd
            LEFT JOIN fact_state_monthly fsm
                ON fsm.state = fsd.state AND fsm.year_month = TO_CHAR(:d, 'YYYY-MM')
            WHERE fsd.fact_date = :d {state_clause}
            ORDER BY fsd.offline_rate DESC
        """), params).fetchall()
    return [dict(r._mapping) for r in rows]


@app.get("/api/dashboard/today-digest")
async def today_digest(
    fact_date: Optional[str] = None,
    user=Depends(verify_token),
    db: Engine = Depends(get_db)
):
    d = date.fromisoformat(fact_date) if fact_date else date.today()
    state_restrict = state_filter(user)
    sr_clause = "AND ff.state = :sr" if state_restrict else ""
    params = {"d": d, "sr": state_restrict}

    with db.connect() as conn:
        # Alert 1: One-visit days
        one_visit = conn.execute(text(f"""
            SELECT ff.employee_code AS employee_id, ff.employee_name, ff.state,
                   ff.data_evidence->>'visits' as visits,
                   ff.data_evidence->>'area_backlog' as backlog,
                   ff.data_evidence->>'state_avg_visits' as state_avg,
                   ff.severity
            FROM fraud_flag ff
            WHERE ff.flag_date = :d AND ff.pattern_id = 'P1' {sr_clause}
            ORDER BY ff.severity DESC, (ff.data_evidence->>'area_backlog')::int DESC
            LIMIT 20
        """), params).fetchall()

        # Alert 3: Chronic PENDING
        chronic_pending = conn.execute(text(f"""
            SELECT ff.ticket_id, ff.employee_name, ff.cs_id,
                   ff.data_evidence->>'reason' as reason,
                   (ff.data_evidence->>'days_reason_unchanged')::text as days_unchanged,
                   ff.state
            FROM fraud_flag ff
            WHERE ff.flag_date = :d AND ff.pattern_id = 'P3' {sr_clause}
            LIMIT 20
        """), params).fetchall()

        # Alert 5: Offline 4+ days no ticket
        no_ticket = conn.execute(text(f"""
            SELECT ff.cs_id, ff.state,
                   ff.data_evidence->>'offline_days' as offline_days,
                   ff.employee_name, ff.employee_code AS employee_id
            FROM fraud_flag ff
            WHERE ff.flag_date = :d AND ff.pattern_id = 'P7' {sr_clause}
            ORDER BY (ff.data_evidence->>'offline_days')::int DESC
            LIMIT 20
        """), params).fetchall()

    return {
        "date": str(d),
        "alert_1_one_visit": [dict(r._mapping) for r in one_visit],
        "alert_3_chronic_pending": [dict(r._mapping) for r in chronic_pending],
        "alert_5_no_ticket_offline": [dict(r._mapping) for r in no_ticket],
    }


@app.get("/api/dashboard/offline-buckets")
async def offline_buckets(
    fact_date: Optional[str] = None,
    user=Depends(verify_token),
    db: Engine = Depends(get_db)
):
    d = date.fromisoformat(fact_date) if fact_date else date.today()
    state_restrict = state_filter(user)
    params = {"d": d, "sr": state_restrict}
    sc = "AND state = :sr" if state_restrict else ""

    with db.connect() as conn:
        r = conn.execute(text(f"""
            SELECT
                SUM(sites_bucket_1_3) as b1_3,
                SUM(sites_bucket_4_7) as b4_7,
                SUM(sites_bucket_8_15) as b8_15,
                SUM(sites_bucket_16_30) as b16_30,
                SUM(sites_bucket_31_60) as b31_60,
                SUM(sites_bucket_61_90) as b61_90,
                SUM(sites_bucket_90plus) as b90plus
            FROM fact_state_daily WHERE fact_date = :d {sc}
        """), params).fetchone()
    return {
        "1-3 Days": int(r.b1_3 or 0),
        "4-7 Days": int(r.b4_7 or 0),
        "8-15 Days": int(r.b8_15 or 0),
        "16-30 Days": int(r.b16_30 or 0),
        "31-60 Days": int(r.b31_60 or 0),
        "61-90 Days": int(r.b61_90 or 0),
        "90+ Days": int(r.b90plus or 0),
    }


@app.get("/api/dashboard/distribution")
async def distribution_graph(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    state: Optional[str] = None,
    metric: str = "visits",
    user=Depends(verify_token),
    db: Engine = Depends(get_db)
):
    """
    Distribution graph data: activity over time.
    
    Metrics: 'visits', 'tickets', 'offline_sites'
    Filters: date range, state (specific or PAN India)
    """
    state_restrict = state_filter(user) or state
    
    if not end_date:
        end_date = date.today()
    else:
        end_date = date.fromisoformat(end_date)
    
    if not start_date:
        start_date = end_date - timedelta(days=30)
    else:
        start_date = date.fromisoformat(start_date)

    # Validate metric
    valid_metrics = ["visits", "tickets", "offline_sites"]
    if metric not in valid_metrics:
        metric = "visits"

    state_filter_sql = "AND vt.site_state_master = :state_restrict" if state_restrict else ""

    with db.connect() as conn:
        if metric == "visits":
            sql = f"""
            SELECT 
                DATE(vt.last_visit_in) as visit_date,
                COUNT(DISTINCT vt.cs_id) as count_value
            FROM view_ticket vt
            WHERE DATE(vt.last_visit_in) BETWEEN :start AND :end
              AND vt.assigned_type = 'ENGINEER'
              {state_filter_sql}
            GROUP BY DATE(vt.last_visit_in)
            ORDER BY visit_date
            """
        
        elif metric == "tickets":
            sql = f"""
            SELECT 
                DATE(vt.create_date) as visit_date,
                COUNT(vt.ticket_id) as count_value
            FROM view_ticket vt
            WHERE DATE(vt.create_date) BETWEEN :start AND :end
              AND vt.assigned_type = 'ENGINEER'
              {state_filter_sql}
            GROUP BY DATE(vt.create_date)
            ORDER BY visit_date
            """
        
        else:  # offline_sites
            sql = f"""
            SELECT 
                odm.data_date as visit_date,
                COUNT(DISTINCT odm.cs_id) as count_value
            FROM offline_data_master odm
            WHERE odm.data_date BETWEEN :start AND :end
              AND odm.segment = 'PSU'
              AND odm.aging_days > 2
              {f"AND odm.state = :state_restrict" if state_restrict else ""}
            GROUP BY odm.data_date
            ORDER BY visit_date
            """
        
        params = {
            "start": start_date,
            "end": end_date,
            "state_restrict": state_restrict
        }
        
        rows = conn.execute(text(sql), params).fetchall()

    # Format data for chart
    data = [{"date": str(r.visit_date), "value": int(r.count_value)} for r in rows]
    
    return {
        "metric": metric,
        "date_range": {"start": str(start_date), "end": str(end_date)},
        "state": state_restrict or "PAN India",
        "data": data,
        "total": sum(d["value"] for d in data) if data else 0
    }


# ─── ENGINEER ENDPOINTS ───────────────────────────────────
@app.get("/api/engineers/leaderboard")
async def engineer_leaderboard(
    year_month: Optional[str] = None,
    state: Optional[str] = None,
    sort_by: str = "composite_score",
    sort_dir: str = "asc",
    page: int = 1,
    page_size: int = 25,
    user=Depends(verify_token),
    db: Engine = Depends(get_db)
):
    ym = year_month or date.today().strftime("%Y-%m")
    state_restrict = state_filter(user) or state

    with db.connect() as conn:
        sc = "AND state = :sr" if state_restrict else ""
        params = {"ym": ym, "sr": state_restrict,
                  "offset": (page - 1) * page_size, "limit": page_size}

        allowed_sorts = ["composite_score","closure_rate","coverage_rate","attendance_rate",
                         "avg_resolution_days","repeat_visit_rate","first_time_fix_rate"]
        if sort_by not in allowed_sorts: sort_by = "composite_score"
        order = "ASC" if sort_dir == "asc" else "DESC"

        rows = conn.execute(text(f"""
            SELECT employee_code AS employee_id, employee_name, state, service_area_code,
                   composite_score, score_band, is_forced_red,
                   closure_rate, coverage_rate, attendance_rate, productive_days_rate,
                   one_visit_days, present_days,
                   avg_resolution_days, repeat_visit_rate, sentback_rate,
                   total_visits, fv_high_flags, pattern_critical_flags,
                   total_tickets, closed_tickets,
                   pop_total_sites, unique_sites_visited
            FROM fact_engineer_monthly
            WHERE year_month = :ym {sc}
            ORDER BY {sort_by} {order}
            LIMIT :limit OFFSET :offset
        """), params).fetchall()

        total = conn.execute(text(f"""
            SELECT COUNT(*) FROM fact_engineer_monthly WHERE year_month = :ym {sc}
        """), params).scalar()

    return {
        "year_month": ym,
        "total": int(total or 0),
        "page": page,
        "page_size": page_size,
        "engineers": [dict(r._mapping) for r in rows]
    }


@app.get("/api/engineers/performance/table")
async def engineer_performance_table(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    state: Optional[str] = None,
    sort_by: str = "employee_name",
    sort_dir: str = "asc",
    page: int = 1,
    page_size: int = 50,
    user=Depends(verify_token),
    db: Engine = Depends(get_db)
):
    """
    Detailed engineer performance metrics table.
    
    Metrics:
    - Service Area, Name, Employee Code, Phone, State
    - Att. Days, Working Days, Att. %
    - Prod. Days, Zero Prod Days
    - Total Visits, Distinct Sites, Repeat Rate (x)
    - Closed, Open, Pending, Complete (tickets)
    - Offline Sites (aging_days > 2 AND segment = 'PSU')
    - Visits Efficiency: (total_visits / total_engineers) / working_days
    """
    state_restrict = state_filter(user) or state
    
    # Default to last 30 days
    if not end_date:
        end_date = date.today()
    else:
        end_date = date.fromisoformat(end_date)
    
    if not start_date:
        start_date = end_date - timedelta(days=30)
    else:
        start_date = date.fromisoformat(start_date)

    with db.connect() as conn:
        # Working days: exclude Sundays + bank holidays
        sql_working_days = """
        SELECT COUNT(*)::INTEGER as working_days
        FROM (
            SELECT GENERATE_SERIES(CAST(:start AS DATE), CAST(:end AS DATE), '1 day'::INTERVAL)::DATE as d
        ) dates
        WHERE EXTRACT(DOW FROM d) != 0
          AND d NOT IN (SELECT holiday_date FROM bank_holiday)
        """
        working_days_result = conn.execute(text(sql_working_days), {"start": start_date, "end": end_date}).fetchone()
        working_days = working_days_result.working_days if working_days_result else 1

        state_filter_sql = "AND em.service_state = :state_restrict" if state_restrict else ""

        # Main query: all engineers with their performance metrics
        sql = f"""
        WITH date_range AS (
            SELECT :start::DATE as start_date, :end::DATE as end_date
        ),
        active_engineers AS (
            SELECT DISTINCT employee_code AS employee_id, employee_name, service_area_code, service_state, phone
            FROM engineer_master
            WHERE UPPER(active_status) = 'YES' {state_filter_sql}
        ),
        attendance_metrics AS (
            SELECT
                ae.employee_id,
                COUNT(DISTINCT a.attendance_date) as att_days
            FROM active_engineers ae
            LEFT JOIN attendance_data a ON ae.employee_id = a.employee_id
                AND a.attendance_date BETWEEN (SELECT start_date FROM date_range) AND (SELECT end_date FROM date_range)
            GROUP BY ae.employee_id
        ),
        visit_metrics AS (
            SELECT
                ae.employee_id,
                COUNT(DISTINCT v.visit_id)        as total_visits,
                COUNT(DISTINCT vt.oracle_site_no) as distinct_sites,
                COUNT(DISTINCT v.visit_date)      as prod_days
            FROM active_engineers ae
            LEFT JOIN visit_master v ON v.employee_id = ae.employee_id
                AND v.visit_date BETWEEN (SELECT start_date FROM date_range) AND (SELECT end_date FROM date_range)
            LEFT JOIN view_ticket vt ON vt.ticket_id = v.ticket_id
            GROUP BY ae.employee_id
        ),
        ticket_metrics AS (
            SELECT
                ae.employee_id,
                COUNT(CASE WHEN vt.ticket_status = 'CLOSED'                        THEN 1 END) as closed_count,
                COUNT(CASE WHEN vt.ticket_status = 'OPEN'                          THEN 1 END) as open_count,
                COUNT(CASE WHEN vt.ticket_status IN ('PENDING', 'SENTBACK')        THEN 1 END) as pending_count,
                COUNT(CASE WHEN vt.ticket_status = 'COMPLETED'                     THEN 1 END) as complete_count
            FROM active_engineers ae
            LEFT JOIN view_ticket vt ON ae.employee_id = vt.assigned_employee_code
            GROUP BY ae.employee_id
        ),
        offline_sites AS (
            SELECT
                em.employee_id,
                COUNT(DISTINCT odm.cs_id) as offline_sites_count
            FROM active_engineers em
            LEFT JOIN customer_site_master csm ON csm.service_area_code = em.service_area_code
            LEFT JOIN offline_data_master odm ON odm.cs_id = csm.cs_id
                AND odm.data_date = (SELECT MAX(data_date) FROM offline_data_master)
                AND odm.aging_days > 2 AND odm.segment = 'PSU'
            GROUP BY em.employee_id
        ),
        service_areas AS (
            SELECT service_area_code, service_area_name FROM service_area_master
        )
        SELECT
            ae.employee_id,
            ae.employee_name,
            ae.service_area_code,
            COALESCE(sam.service_area_name, ae.service_area_code) as service_area_name,
            ae.service_state,
            ae.phone,
            COALESCE(am.att_days, 0)  as att_days,
            :working_days::INTEGER    as working_days,
            CASE WHEN :working_days > 0 THEN ROUND(COALESCE(am.att_days, 0)::NUMERIC / :working_days * 100, 2) ELSE 0 END as att_percentage,
            COALESCE(vm.prod_days, 0) as prod_days,
            COALESCE(am.att_days, 0) - COALESCE(vm.prod_days, 0) as zero_prod_days,
            COALESCE(vm.total_visits, 0)    as total_visits,
            COALESCE(vm.distinct_sites, 0)  as distinct_sites,
            CASE WHEN COALESCE(vm.distinct_sites, 0) > 0
                 THEN ROUND(COALESCE(vm.total_visits, 0)::NUMERIC / COALESCE(vm.distinct_sites, 1), 2)
                 ELSE 0 END as repeat_rate,
            COALESCE(tm.closed_count, 0)    as closed_tickets,
            COALESCE(tm.open_count, 0)      as open_tickets,
            COALESCE(tm.pending_count, 0)   as pending_tickets,
            COALESCE(tm.complete_count, 0)  as completed_tickets,
            COALESCE(os.offline_sites_count, 0) as offline_sites
        FROM active_engineers ae
        LEFT JOIN attendance_metrics am ON ae.employee_id = am.employee_id
        LEFT JOIN visit_metrics vm      ON ae.employee_id = vm.employee_id
        LEFT JOIN ticket_metrics tm     ON ae.employee_id = tm.employee_id
        LEFT JOIN offline_sites os      ON ae.employee_id = os.employee_id
        LEFT JOIN service_areas sam     ON ae.service_area_code = sam.service_area_code
        ORDER BY {sort_by} {sort_dir.upper()}
        LIMIT :limit OFFSET :offset
        """
        
        params = {
            "start": start_date,
            "end": end_date,
            "working_days": working_days,
            "state_restrict": state_restrict,
            "offset": (page - 1) * page_size,
            "limit": page_size
        }
        
        # Validate sort_by
        allowed_sorts = [
            "employee_name", "service_area_name", "service_state", "att_days", "att_percentage",
            "prod_days", "total_visits", "distinct_sites", "repeat_rate",
            "closed_tickets", "open_tickets", "pending_tickets", "offline_sites"
        ]
        if sort_by not in allowed_sorts:
            sort_by = "employee_name"
            sql = sql.replace("{sort_by}", sort_by)
        
        rows = conn.execute(text(sql), params).fetchall()
        
        # Get total count
        count_sql = f"""
        SELECT COUNT(*)
        FROM engineer_master
        WHERE UPPER(is_active) = 'YES' {state_filter_sql}
        """
        total = conn.execute(text(count_sql), {"state_restrict": state_restrict}).scalar() or 0

    # Calculate visits efficiency metric
    total_engineers = max(int(total), 1)
    total_visits_all = sum(r.total_visits for r in rows) if rows else 0
    visits_efficiency = (total_visits_all / total_engineers) / max(working_days, 1)

    return {
        "date_range": {"start": str(start_date), "end": str(end_date)},
        "state": state_restrict or "PAN India",
        "working_days": working_days,
        "total_engineers": total_engineers,
        "visits_efficiency_metric": round(visits_efficiency, 3),
        "page": page,
        "page_size": page_size,
        "total": int(total),
        "engineers": [dict(r._mapping) for r in rows]
    }


@app.get("/api/engineers/performance/export")
async def export_performance_table(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    state: Optional[str] = None,
    user=Depends(verify_token),
    db: Engine = Depends(get_db)
):
    """
    Export engineer performance table to Excel.
    Fetches all data (no pagination) and returns as file download.
    """
    state_restrict = state_filter(user) or state
    
    if not end_date:
        end_date = date.today()
    else:
        end_date = date.fromisoformat(end_date)
    
    if not start_date:
        start_date = end_date - timedelta(days=30)
    else:
        start_date = date.fromisoformat(start_date)

    with db.connect() as conn:
        # Working days: exclude Sundays + bank holidays
        sql_working_days = """
        SELECT COUNT(*)::INTEGER as working_days
        FROM (
            SELECT GENERATE_SERIES(CAST(:start AS DATE), CAST(:end AS DATE), '1 day'::INTERVAL)::DATE as d
        ) dates
        WHERE EXTRACT(DOW FROM d) != 0
          AND d NOT IN (SELECT holiday_date FROM bank_holiday)
        """
        working_days_result = conn.execute(text(sql_working_days), {"start": start_date, "end": end_date}).fetchone()
        working_days = working_days_result.working_days if working_days_result else 1

        state_filter_sql = "AND em.service_state = :state_restrict" if state_restrict else ""

        # Same query as performance table but without LIMIT/OFFSET
        sql = f"""
        WITH date_range AS (
            SELECT :start::DATE as start_date, :end::DATE as end_date
        ),
        active_engineers AS (
            SELECT DISTINCT employee_code AS employee_id, employee_name, service_area_code, service_state, phone
            FROM engineer_master
            WHERE UPPER(active_status) = 'YES' {state_filter_sql}
        ),
        attendance_metrics AS (
            SELECT
                ae.employee_id,
                COUNT(DISTINCT a.attendance_date) as att_days
            FROM active_engineers ae
            LEFT JOIN attendance_data a ON ae.employee_id = a.employee_id
                AND a.attendance_date BETWEEN (SELECT start_date FROM date_range) AND (SELECT end_date FROM date_range)
            GROUP BY ae.employee_id
        ),
        visit_metrics AS (
            SELECT
                ae.employee_id,
                COUNT(DISTINCT v.visit_id)        as total_visits,
                COUNT(DISTINCT vt.oracle_site_no) as distinct_sites,
                COUNT(DISTINCT v.visit_date)      as prod_days
            FROM active_engineers ae
            LEFT JOIN visit_master v ON v.employee_id = ae.employee_id
                AND v.visit_date BETWEEN (SELECT start_date FROM date_range) AND (SELECT end_date FROM date_range)
            LEFT JOIN view_ticket vt ON vt.ticket_id = v.ticket_id
            GROUP BY ae.employee_id
        ),
        ticket_metrics AS (
            SELECT
                ae.employee_id,
                COUNT(CASE WHEN vt.ticket_status = 'CLOSED'                 THEN 1 END) as closed_count,
                COUNT(CASE WHEN vt.ticket_status = 'OPEN'                   THEN 1 END) as open_count,
                COUNT(CASE WHEN vt.ticket_status IN ('PENDING', 'SENTBACK') THEN 1 END) as pending_count,
                COUNT(CASE WHEN vt.ticket_status = 'COMPLETED'              THEN 1 END) as complete_count
            FROM active_engineers ae
            LEFT JOIN view_ticket vt ON ae.employee_id = vt.assigned_employee_code
            GROUP BY ae.employee_id
        ),
        offline_sites AS (
            SELECT
                em.employee_id,
                COUNT(DISTINCT odm.cs_id) as offline_sites_count
            FROM active_engineers em
            LEFT JOIN customer_site_master csm ON csm.service_area_code = em.service_area_code
            LEFT JOIN offline_data_master odm ON odm.cs_id = csm.cs_id
                AND odm.data_date = (SELECT MAX(data_date) FROM offline_data_master)
                AND odm.aging_days > 2 AND odm.segment = 'PSU'
            GROUP BY em.employee_id
        ),
        service_areas AS (
            SELECT service_area_code, service_area_name FROM service_area_master
        )
        SELECT
            ae.employee_id,
            ae.employee_name,
            ae.service_area_code,
            COALESCE(sam.service_area_name, ae.service_area_code) as service_area_name,
            ae.service_state,
            ae.phone,
            COALESCE(am.att_days, 0)  as att_days,
            :working_days::INTEGER    as working_days,
            CASE WHEN :working_days > 0 THEN ROUND(COALESCE(am.att_days, 0)::NUMERIC / :working_days * 100, 2) ELSE 0 END as att_percentage,
            COALESCE(vm.prod_days, 0) as prod_days,
            COALESCE(am.att_days, 0) - COALESCE(vm.prod_days, 0) as zero_prod_days,
            COALESCE(vm.total_visits, 0)   as total_visits,
            COALESCE(vm.distinct_sites, 0) as distinct_sites,
            CASE WHEN COALESCE(vm.distinct_sites, 0) > 0
                 THEN ROUND(COALESCE(vm.total_visits, 0)::NUMERIC / COALESCE(vm.distinct_sites, 1), 2)
                 ELSE 0 END as repeat_rate,
            COALESCE(tm.closed_count, 0)   as closed_tickets,
            COALESCE(tm.open_count, 0)     as open_tickets,
            COALESCE(tm.pending_count, 0)  as pending_tickets,
            COALESCE(tm.complete_count, 0) as completed_tickets,
            COALESCE(os.offline_sites_count, 0) as offline_sites
        FROM active_engineers ae
        LEFT JOIN attendance_metrics am ON ae.employee_id = am.employee_id
        LEFT JOIN visit_metrics vm      ON ae.employee_id = vm.employee_id
        LEFT JOIN ticket_metrics tm     ON ae.employee_id = tm.employee_id
        LEFT JOIN offline_sites os      ON ae.employee_id = os.employee_id
        LEFT JOIN service_areas sam     ON ae.service_area_code = sam.service_area_code
        ORDER BY ae.employee_name
        """
        
        params = {
            "start": start_date,
            "end": end_date,
            "working_days": working_days,
            "state_restrict": state_restrict
        }
        
        rows = conn.execute(text(sql), params).fetchall()

    # Convert to pandas DataFrame and create Excel
    df = pd.DataFrame([dict(r._mapping) for r in rows])
    
    if df.empty:
        raise HTTPException(status_code=400, detail="No data to export")

    # Reorder columns for better readability
    column_order = [
        'employee_name', 'employee_id', 'phone', 'service_area_name', 'service_state',
        'att_days', 'working_days', 'att_percentage', 'prod_days', 'zero_prod_days',
        'total_visits', 'distinct_sites', 'repeat_rate',
        'closed_tickets', 'open_tickets', 'pending_tickets', 'completed_tickets',
        'offline_sites'
    ]
    
    existing_cols = [c for c in column_order if c in df.columns]
    df = df[existing_cols]

    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Performance', index=False)
        
        # Format headers
        from openpyxl.styles import Font, PatternFill, Alignment
        ws = writer.sheets['Performance']
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-width columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    output.seek(0)
    filename = f"engineer_performance_{start_date}_{end_date}.xlsx"
    return FileResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename
    )


@app.get("/api/engineers/{employee_code}/profile")
async def engineer_profile(
    employee_code: str,
    year_month: Optional[str] = None,
    user=Depends(verify_token),
    db: Engine = Depends(get_db)
):
    ym = year_month or date.today().strftime("%Y-%m")
    state_restrict = state_filter(user)

    with db.connect() as conn:
        # Basic info
        eng = conn.execute(text("""
            SELECT em.employee_code AS employee_id, em.employee_name, em.service_state, em.service_area_code,
                   em.designation,
                   fem.composite_score, fem.score_band, fem.is_forced_red,
                   fem.closure_rate, fem.coverage_rate, fem.attendance_rate,
                   fem.productive_days, fem.present_days, fem.one_visit_days,
                   fem.total_visits, fem.repeat_visit_rate,
                   fem.avg_resolution_days, fem.sentback_rate,
                   fem.pop_total_sites, fem.unique_sites_visited,
                   fem.fv_high_flags, fem.pattern_critical_flags,
                   fem.score_closure, fem.score_coverage, fem.score_resolution,
                   fem.score_area_offline, fem.score_repeat, fem.score_attendance, fem.score_fraud
            FROM engineer_master em
            LEFT JOIN fact_engineer_monthly fem
                ON fem.employee_code = em.employee_code AND fem.year_month = :ym
            WHERE em.employee_code = :ec
        """), {"ec": employee_code, "ym": ym}).fetchone()

        if not eng:
            raise HTTPException(404, "Engineer not found")

        if state_restrict and eng.service_state != state_restrict:
            raise HTTPException(403, "Access denied")

        # Monthly trend (last 4 months)
        trend = conn.execute(text("""
            SELECT year_month, composite_score, closure_rate, attendance_rate,
                   productive_days, present_days
            FROM fact_engineer_monthly
            WHERE employee_code = :ec
            ORDER BY year_month DESC LIMIT 4
        """), {"ec": employee_code}).fetchall()

        # Daily attendance calendar (current month)
        ym_start = datetime.strptime(ym + "-01", "%Y-%m-%d").date()
        ym_end = (ym_start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        calendar = conn.execute(text("""
            SELECT a.attendance_date, a.attendance_status_derived AS attendance_status,
                   a.in_datetime,
                   fed.daily_visits, fed.daily_closures, fed.is_productive_day,
                   fed.is_one_visit_day, fed.state_avg_visits
            FROM attendance_data a
            LEFT JOIN fact_engineer_daily fed
                ON fed.employee_code = a.employee_id AND fed.fact_date = a.attendance_date
            WHERE a.employee_id = :ec
              AND a.attendance_date BETWEEN :s AND :e
            ORDER BY a.attendance_date
        """), {"ec": employee_code, "s": ym_start, "e": ym_end}).fetchall()

        # Fraud flags
        flags = conn.execute(text("""
            SELECT pattern_id, severity, description, flag_date,
                   ticket_id, cs_id, review_status, data_evidence
            FROM fraud_flag
            WHERE employee_code = :ec AND flag_date >= :s
            ORDER BY flag_date DESC LIMIT 50
        """), {"ec": employee_code, "s": ym_start}).fetchall()

        # PoP sites
        pop_sites = conn.execute(text("""
            SELECT csm.cs_id, csm.site_name, csm.state,
                   odm.aging_days, odm.bucket,
                   vt.ticket_status, vt.last_visit_in,
                   vt.ticket_aging_days
            FROM customer_site_master csm
            LEFT JOIN offline_data_master odm
                ON odm.cs_id = csm.cs_id
                AND odm.data_date = (SELECT MAX(data_date) FROM offline_data_master)
            LEFT JOIN view_ticket vt ON vt.cs_id = csm.cs_id
                AND vt.ticket_status IN ('OPEN','PENDING','SENTBACK','COMPLETED')
            WHERE csm.service_area_code = :sac
            ORDER BY odm.aging_days DESC NULLS LAST
            LIMIT 100
        """), {"sac": eng.service_area_code}).fetchall()

    return {
        "engineer": dict(eng._mapping),
        "monthly_trend": [dict(r._mapping) for r in trend],
        "attendance_calendar": [dict(r._mapping) for r in calendar],
        "fraud_flags": [dict(r._mapping) for r in flags],
        "pop_sites": [dict(r._mapping) for r in pop_sites],
    }


# ─── TICKET ENDPOINTS ─────────────────────────────────────
@app.get("/api/tickets/analysis")
async def ticket_analysis(
    upload_date: Optional[str] = None,
    state: Optional[str] = None,
    status: Optional[str] = None,
    user=Depends(verify_token),
    db: Engine = Depends(get_db)
):
    d = date.fromisoformat(upload_date) if upload_date else date.today()
    state_restrict = state_filter(user) or state
    params = {"d": d, "sr": state_restrict}
    sc = "AND (vt.site_state_master = :sr OR vt.state_name = :sr)" if state_restrict else ""

    with db.connect() as conn:
        # Status distribution
        status_dist = conn.execute(text(f"""
            SELECT ticket_status, COUNT(*) as count,
                   ROUND(COUNT(*)::numeric / SUM(COUNT(*)) OVER () * 100, 1) as pct
            FROM view_ticket vt
            WHERE assigned_type = 'ENGINEER' {sc}
            GROUP BY ticket_status ORDER BY count DESC
        """), params).fetchall()

        # Aging bands (OPEN + PENDING)
        aging = conn.execute(text(f"""
            SELECT
                COUNT(*) FILTER (WHERE ticket_aging_days <= 1) as d0_1,
                COUNT(*) FILTER (WHERE ticket_aging_days BETWEEN 2 AND 3) as d2_3,
                COUNT(*) FILTER (WHERE ticket_aging_days BETWEEN 4 AND 7) as d4_7,
                COUNT(*) FILTER (WHERE ticket_aging_days BETWEEN 8 AND 15) as d8_15,
                COUNT(*) FILTER (WHERE ticket_aging_days > 15) as d15plus
            FROM view_ticket vt
            WHERE ticket_status IN ('OPEN','PENDING') {sc}
        """), params).fetchone()

        # Breach list (>3 days OPEN or PENDING)
        breach = conn.execute(text(f"""
            SELECT vt.ticket_id, vt.cs_id, vt.state_name, vt.ticket_status,
                   vt.ticket_aging_days, vt.planned_date, vt.ticket_status_reason,
                   vt.assigned_employee_code, vt.assigned_to,
                   csm.site_name
            FROM view_ticket vt
            LEFT JOIN customer_site_master csm ON csm.cs_id = vt.cs_id
            WHERE vt.ticket_status IN ('OPEN','PENDING')
              AND vt.ticket_aging_days > 3 {sc}
            ORDER BY vt.ticket_aging_days DESC
            LIMIT 100
        """), params).fetchall()

        # SENTBACK analysis by engineer
        sentback = conn.execute(text(f"""
            SELECT vt.assigned_employee_code, vt.assigned_to,
                   COUNT(*) as sentback_count
            FROM view_ticket vt
            WHERE vt.ticket_status = 'SENTBACK' {sc}
            GROUP BY vt.assigned_employee_code, vt.assigned_to
            ORDER BY sentback_count DESC LIMIT 20
        """), params).fetchall()

        # No-ticket offline sites (latest offline data)
        latest_offline_d = conn.execute(text("SELECT MAX(data_date) FROM offline_data_master")).scalar()
        no_ticket_sites = conn.execute(text(f"""
            SELECT odm.cs_id, odm.site_name, odm.state, odm.aging_days,
                   odm.bucket, odm.b2b_code, csm.service_area_code,
                   em.employee_name, em.employee_code AS employee_id
            FROM offline_data_master odm
            LEFT JOIN customer_site_master csm ON csm.cs_id = odm.cs_id
            LEFT JOIN engineer_master em ON em.service_area_code = csm.service_area_code
            WHERE odm.data_date = :latest_d AND odm.aging_days >= 4 AND odm.segment = 'PSU'
              AND NOT EXISTS (
                  SELECT 1 FROM view_ticket vt
                  WHERE vt.cs_id = odm.cs_id
                    AND vt.ticket_status IN ('OPEN','PENDING','SENTBACK','COMPLETED')
              )
            {'AND odm.state = :sr' if state_restrict else ''}
            ORDER BY odm.aging_days DESC
            LIMIT 100
        """), {**params, "latest_d": latest_offline_d}).fetchall()

    return {
        "date": str(d),
        "status_distribution": [dict(r._mapping) for r in status_dist],
        "aging_bands": dict(aging._mapping) if aging else {},
        "breach_list": [dict(r._mapping) for r in breach],
        "sentback_by_engineer": [dict(r._mapping) for r in sentback],
        "no_ticket_offline_sites": [dict(r._mapping) for r in no_ticket_sites],
    }


@app.get("/api/tickets/chronic-pending")
async def chronic_pending(
    analysis_date: Optional[str] = None,
    user=Depends(verify_token),
    db: Engine = Depends(get_db)
):
    d = date.fromisoformat(analysis_date) if analysis_date else date.today()
    state_restrict = state_filter(user)
    sc = "AND ff.state = :sr" if state_restrict else ""

    with db.connect() as conn:
        rows = conn.execute(text(f"""
            SELECT ff.ticket_id, ff.cs_id, ff.employee_code AS employee_id, ff.employee_name,
                   ff.state, ff.description, ff.data_evidence,
                   ff.severity, ff.flag_date
            FROM fraud_flag ff
            WHERE ff.flag_date = :d AND ff.pattern_id = 'P3' {sc}
            ORDER BY (ff.data_evidence->>'aging_days')::int DESC
        """), {"d": d, "sr": state_restrict}).fetchall()
    return [dict(r._mapping) for r in rows]


# ─── FRAUD ENDPOINTS ──────────────────────────────────────
@app.get("/api/fraud/flags")
async def fraud_flags(
    flag_date: Optional[str] = None,
    pattern_id: Optional[str] = None,
    severity: Optional[str] = None,
    review_status: str = "PENDING",
    state: Optional[str] = None,
    user=Depends(verify_token),
    db: Engine = Depends(get_db)
):
    d = date.fromisoformat(flag_date) if flag_date else date.today()
    state_restrict = state_filter(user) or state
    params = {"d": d, "sr": state_restrict, "pid": pattern_id,
              "sv": severity, "rs": review_status}

    filters = "WHERE ff.flag_date = :d"
    if state_restrict: filters += " AND ff.state = :sr"
    if pattern_id: filters += " AND ff.pattern_id = :pid"
    if severity: filters += " AND ff.severity = :sv"
    if review_status != "ALL": filters += " AND ff.review_status = :rs"

    with db.connect() as conn:
        rows = conn.execute(text(f"""
            SELECT ff.id, ff.flag_date, ff.employee_code AS employee_id, ff.employee_name,
                   ff.state, ff.ticket_id, ff.cs_id,
                   ff.pattern_type, ff.pattern_id, ff.severity,
                   ff.description, ff.data_evidence, ff.review_status,
                   ff.score_impact, ff.created_at
            FROM fraud_flag ff
            {filters}
            ORDER BY CASE ff.severity WHEN 'CRITICAL' THEN 1 WHEN 'HIGH' THEN 2 ELSE 3 END,
                     ff.flag_date DESC
            LIMIT 200
        """), params).fetchall()
    return [dict(r._mapping) for r in rows]


@app.post("/api/fraud/flags/{flag_id}/review")
async def review_fraud_flag(
    flag_id: int,
    action: str = Form(...),   # 'VALID' or 'CONFIRMED_FAKE'
    note: str = Form(""),
    user=Depends(require_ops_or_admin),
    db: Engine = Depends(get_db)
):
    if action not in ("VALID", "CONFIRMED_FAKE"):
        raise HTTPException(400, "action must be VALID or CONFIRMED_FAKE")

    with db.begin() as conn:
        conn.execute(text("""
            UPDATE fraud_flag
            SET review_status = :action, reviewed_by = :by,
                reviewed_at = NOW(), review_note = :note
            WHERE id = :id
        """), {"action": action, "by": user["username"], "note": note, "id": flag_id})
    return {"message": f"Flag {flag_id} marked as {action}"}


# ─── STATE ENDPOINTS ──────────────────────────────────────
@app.get("/api/states/comparison")
async def state_comparison(
    year_month: Optional[str] = None,
    user=Depends(require_ops_or_admin),
    db: Engine = Depends(get_db)
):
    ym = year_month or date.today().strftime("%Y-%m")
    with db.connect() as conn:
        rows = conn.execute(text("""
            SELECT fsm.state, fsm.zone, fsm.state_score, fsm.score_band,
                   fsm.total_engineers, fsm.avg_engineer_score,
                   fsm.avg_offline_rate, fsm.sites_offline_30plus,
                   fsm.closure_rate, fsm.avg_resolution_days,
                   fsm.carryforward_sites, fsm.carryforward_rate,
                   fsm.fraud_flags_count, fsm.productive_days_rate,
                   fsd.offline_sites, fsd.total_sites, fsd.open_tickets
            FROM fact_state_monthly fsm
            LEFT JOIN fact_state_daily fsd
                ON fsd.state = fsm.state AND fsd.fact_date = CURRENT_DATE
            WHERE fsm.year_month = :ym
            ORDER BY fsm.state_score ASC
        """), {"ym": ym}).fetchall()
    return [dict(r._mapping) for r in rows]


# ─── USER MANAGEMENT ──────────────────────────────────────
@app.get("/api/users")
async def list_users(user=Depends(require_ops_or_admin), db: Engine = Depends(get_db)):
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    with db.connect() as conn:
        rows = conn.execute(text("""
            SELECT id, username, role, state_code, is_active, last_login, created_at
            FROM app_user ORDER BY created_at DESC
        """)).fetchall()
    return [dict(r._mapping) for r in rows]


class UserCreate(BaseModel):
    username: str
    password: str
    role: str
    state_code: Optional[str] = None


@app.post("/api/users")
async def create_user(
    data: UserCreate,
    user=Depends(require_ops_or_admin),
    db: Engine = Depends(get_db)
):
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    if len(data.password) < 10:
        raise HTTPException(400, "Password must be at least 10 characters")
    pw_hash = bcrypt.hashpw(data.password.encode(), bcrypt.gensalt()).decode()
    with db.begin() as conn:
        conn.execute(text("""
            INSERT INTO app_user (username, password_hash, role, state_code, created_by)
            VALUES (:u, :ph, :r, :sc, :cb)
        """), {"u": data.username, "ph": pw_hash, "r": data.role,
               "sc": data.state_code, "cb": user["username"]})
    return {"message": f"User {data.username} created"}


# ─── SETTINGS ─────────────────────────────────────────────
@app.get("/api/settings")
async def get_settings(user=Depends(require_ops_or_admin), db: Engine = Depends(get_db)):
    with db.connect() as conn:
        rows = conn.execute(text("SELECT key, value, description FROM app_settings")).fetchall()
    return {r.key: {"value": r.value, "description": r.description} for r in rows}


@app.put("/api/settings/{key}")
async def update_setting(
    key: str, value: str = Form(...),
    user=Depends(require_ops_or_admin), db: Engine = Depends(get_db)
):
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    with db.begin() as conn:
        conn.execute(text("""
            UPDATE app_settings SET value = :v, updated_at = NOW() WHERE key = :k
        """), {"v": value, "k": key})
    return {"message": f"Setting {key} updated"}


# ─── HEALTH CHECK ─────────────────────────────────────────
@app.get("/api/health")
async def health(db: Engine = Depends(get_db)):
    try:
        with db.connect() as conn:
            conn.execute(text("SELECT 1"))
        with db.connect() as conn:
            last_upload = conn.execute(text(
                "SELECT value FROM app_settings WHERE key = 'last_etl_run'"
            )).scalar()
        return {"status": "healthy", "last_etl_run": last_upload}
    except Exception as e:
        return JSONResponse(status_code=503, content={"status": "unhealthy", "error": str(e)})


# ─── SERVE FRONTEND ───────────────────────────────────────
frontend_path = Path(__file__).parent.parent.parent / "frontend" / "dist"
if frontend_path.exists():
    app.mount("/", StaticFiles(directory=str(frontend_path), html=True), name="static")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)


# ─── DAILY OPERATIONS REPORT ──────────────────────────────

@app.get("/api/ops/sites-visited-datewise")
async def sites_visited_datewise(user=Depends(verify_token), db: Engine = Depends(get_db)):
    state_restrict = state_filter(user)
    rows = []
    with db.connect() as conn:
        for i in range(7):
            d = date.today() - timedelta(days=i+1)
            sc = "AND site_state_master = :sr" if state_restrict else ""
            count = conn.execute(text(f"SELECT COUNT(DISTINCT cs_id) FROM view_ticket WHERE DATE(last_visit_in) = :d AND assigned_employee_code != '' AND ticket_assigned_type = 'ENGINEER' {sc}"), {"d": d, "sr": state_restrict}).scalar() or 0
            rows.append({"date": str(d), "day_label": d.strftime("%d-%m-%Y"), "sites_visited": int(count)})
    return list(reversed(rows))


@app.get("/api/ops/engineers-used-app-datewise")
async def engineers_used_app_datewise(user=Depends(verify_token), db: Engine = Depends(get_db)):
    state_restrict = state_filter(user)
    rows = []
    with db.connect() as conn:
        tc = "AND service_state = :sr" if state_restrict else ""
        total_eng = conn.execute(text(f"SELECT COUNT(*) FROM engineer_master WHERE 1=1 {tc}"), {"sr": state_restrict}).scalar() or 1
        for i in range(7):
            d = date.today() - timedelta(days=i+1)
            sc = "AND site_state_master = :sr" if state_restrict else ""
            count = conn.execute(text(f"SELECT COUNT(DISTINCT assigned_employee_code) FROM view_ticket WHERE DATE(last_visit_in) = :d AND assigned_employee_code != '' AND ticket_assigned_type = 'ENGINEER' {sc}"), {"d": d, "sr": state_restrict}).scalar() or 0
            rows.append({"date": str(d), "day_label": d.strftime("%d-%m-%Y"), "engineers_used_app": int(count), "total_engineers": int(total_eng), "percentage": round(int(count) / int(total_eng) * 100, 3) if total_eng > 0 else 0})
    return list(reversed(rows))


@app.get("/api/ops/statewise-visits")
async def statewise_visits(user=Depends(require_ops_or_admin), db: Engine = Depends(get_db)):
    yesterday = date.today() - timedelta(days=1)
    day_before = date.today() - timedelta(days=2)
    rows = []
    grand = {"y": 0, "db": 0, "eng": 0, "cl": 0, "off3": 0}
    with db.connect() as conn:
        # Build offline >3 days map from latest b2b_offline upload (state -> count)
        off3_rows = conn.execute(text("""
            SELECT LOWER(state) as st, COUNT(*) as cnt
            FROM offline_data_master
            WHERE data_date = (SELECT MAX(data_date) FROM offline_data_master)
              AND aging_days > 3 AND segment = 'PSU'
            GROUP BY LOWER(state)
        """)).fetchall()
        off3_map = {r.st: int(r.cnt) for r in off3_rows}

        states = [r.service_state for r in conn.execute(text("SELECT DISTINCT service_state FROM engineer_master WHERE service_state IS NOT NULL AND service_state != '' ORDER BY service_state")).fetchall()]
        for state in states:
            vy = conn.execute(text("SELECT COUNT(DISTINCT vt.cs_id) FROM view_ticket vt JOIN engineer_master em ON em.employee_code=vt.assigned_employee_code WHERE DATE(vt.last_visit_in)=:d AND em.service_state=:st AND vt.ticket_assigned_type='ENGINEER'"), {"d": yesterday, "st": state}).scalar() or 0
            vdb = conn.execute(text("SELECT COUNT(DISTINCT vt.cs_id) FROM view_ticket vt JOIN engineer_master em ON em.employee_code=vt.assigned_employee_code WHERE DATE(vt.last_visit_in)=:d AND em.service_state=:st AND vt.ticket_assigned_type='ENGINEER'"), {"d": day_before, "st": state}).scalar() or 0
            te = conn.execute(text("SELECT COUNT(*) FROM engineer_master WHERE service_state=:st"), {"st": state}).scalar() or 0
            cl = conn.execute(text("SELECT COUNT(*) FROM view_ticket vt JOIN engineer_master em ON em.employee_code=vt.assigned_employee_code WHERE DATE(vt.ticket_closed_datetime)=:d AND em.service_state=:st AND vt.ticket_status='CLOSED' AND vt.ticket_assigned_type='ENGINEER'"), {"d": yesterday, "st": state}).scalar() or 0
            avg_cl = round(int(cl) / int(te), 1) if te > 0 else 0.0
            offline_3d = off3_map.get(state.lower(), 0)
            grand["y"] += int(vy); grand["db"] += int(vdb); grand["eng"] += int(te); grand["cl"] += int(cl); grand["off3"] += offline_3d
            rows.append({"state": state, "visited_day_before": int(vdb), "visited_yesterday": int(vy), "offline_gt3": offline_3d, "total_engineers": int(te), "avg_closure_per_engineer": avg_cl})
        rows.append({"state": "Grand Total", "visited_day_before": grand["db"], "visited_yesterday": grand["y"], "offline_gt3": grand["off3"], "total_engineers": grand["eng"], "avg_closure_per_engineer": round(grand["cl"] / grand["eng"], 1) if grand["eng"] > 0 else 0.0})
    return {"yesterday": str(yesterday), "day_before": str(day_before), "data": rows}


@app.get("/api/ops/attendance-analysis")
async def attendance_analysis(target_date: Optional[str] = None, user=Depends(verify_token), db: Engine = Depends(get_db)):
    d = date.fromisoformat(target_date) if target_date else date.today() - timedelta(days=1)
    state_restrict = state_filter(user)
    rows = []
    grand = {"tot": 0, "pin": 0, "late": 0, "pre": 0, "np": 0}
    with db.connect() as conn:
        sc = "AND em.service_state = :sr" if state_restrict else ""
        states = [r.service_state for r in conn.execute(text(f"SELECT DISTINCT em.service_state FROM engineer_master em WHERE em.service_state IS NOT NULL AND em.service_state != '' {sc} ORDER BY em.service_state"), {"sr": state_restrict}).fetchall()]
        for state in states:
            te = conn.execute(text("SELECT COUNT(*) FROM engineer_master WHERE service_state=:st"), {"st": state}).scalar() or 0
            pin = conn.execute(text("SELECT COUNT(DISTINCT a.employee_id) FROM attendance_data a JOIN engineer_master em ON em.employee_code=a.employee_id WHERE a.attendance_date=:d AND em.service_state=:st AND a.attendance_status_derived IN ('ON TIME','LATE')"), {"d": d, "st": state}).scalar() or 0
            on_time = conn.execute(text("SELECT COUNT(DISTINCT a.employee_id) FROM attendance_data a JOIN engineer_master em ON em.employee_code=a.employee_id WHERE a.attendance_date=:d AND em.service_state=:st AND a.attendance_status_derived='ON TIME'"), {"d": d, "st": state}).scalar() or 0
            late = int(pin) - int(on_time)
            np = int(te) - int(pin)
            pct_pin = round(int(pin)/int(te)*100, 1) if te > 0 else 0.0
            pct_ot = round(int(on_time)/int(te)*100, 1) if te > 0 else 0.0
            grand["tot"] += int(te); grand["pin"] += int(pin); grand["late"] += late; grand["pre"] += int(pin); grand["np"] += np
            rows.append({"state": state, "total_engineers": int(te), "total_punched_in": int(pin), "late": late, "present": int(pin), "not_punched": np, "pct_punched": pct_pin, "pct_on_time": pct_ot})
        rows.append({"state": "Grand Total", "total_engineers": grand["tot"], "total_punched_in": grand["pin"], "late": grand["late"], "present": grand["pre"], "not_punched": grand["np"], "pct_punched": round(grand["pin"]/grand["tot"]*100,1) if grand["tot"]>0 else 0.0, "pct_on_time": round((grand["pin"]-grand["late"])/grand["tot"]*100,1) if grand["tot"]>0 else 0.0})
    return {"date": str(d), "data": rows}


# ─── VISIT FORMS ──────────────────────────────────────────────────────────────

@app.get("/api/visits/analysis")
async def visits_analysis(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    state: Optional[str] = None,
    user=Depends(verify_token),
    db: Engine = Depends(get_db)
):
    state_restrict = state_filter(user) or state
    end_d   = date.fromisoformat(end_date)   if end_date   else date.today()
    start_d = date.fromisoformat(start_date) if start_date else end_d - timedelta(days=29)

    with db.connect() as conn:
        sc = "AND em.service_state = :sr" if state_restrict else ""
        params = {"s": start_d, "e": end_d, "sr": state_restrict}

        # Summary totals
        totals = conn.execute(text(f"""
            SELECT
                COUNT(*)                                                          AS total_visits,
                COUNT(DISTINCT vf.cs_id)                                         AS unique_sites,
                COUNT(DISTINCT vf.employee_code)                                 AS unique_engineers,
                SUM(CASE WHEN UPPER(vf.problem_solved) = 'YES' THEN 1 ELSE 0 END) AS solved,
                SUM(CASE WHEN UPPER(vf.pm_done) = 'YES' THEN 1 ELSE 0 END)       AS pm_done
            FROM visit_form vf
            LEFT JOIN engineer_master em ON em.employee_code = vf.employee_code
            WHERE vf.visit_date BETWEEN :s AND :e {sc}
        """), params).fetchone()

        # Daily trend (last 30 days)
        daily = conn.execute(text(f"""
            SELECT vf.visit_date,
                   COUNT(*)                                                           AS visits,
                   SUM(CASE WHEN UPPER(vf.problem_solved) = 'YES' THEN 1 ELSE 0 END) AS solved,
                   SUM(CASE WHEN UPPER(vf.pm_done) = 'YES' THEN 1 ELSE 0 END)        AS pm_done
            FROM visit_form vf
            LEFT JOIN engineer_master em ON em.employee_code = vf.employee_code
            WHERE vf.visit_date BETWEEN :s AND :e {sc}
            GROUP BY vf.visit_date ORDER BY vf.visit_date
        """), params).fetchall()

        # Ticket status breakdown from visit form
        status_breakdown = conn.execute(text(f"""
            SELECT COALESCE(vf.ticket_status, 'Unknown') AS status, COUNT(*) AS cnt
            FROM visit_form vf
            LEFT JOIN engineer_master em ON em.employee_code = vf.employee_code
            WHERE vf.visit_date BETWEEN :s AND :e {sc}
            GROUP BY vf.ticket_status ORDER BY cnt DESC LIMIT 10
        """), params).fetchall()

        # Top engineers by visits
        top_engs = conn.execute(text(f"""
            SELECT vf.employee_code,
                   COALESCE(em.employee_name, vf.technician_name, vf.employee_code) AS name,
                   COALESCE(em.service_state, '') AS state,
                   COUNT(*)                                                           AS visits,
                   SUM(CASE WHEN UPPER(vf.problem_solved) = 'YES' THEN 1 ELSE 0 END) AS solved,
                   SUM(CASE WHEN UPPER(vf.pm_done) = 'YES' THEN 1 ELSE 0 END)        AS pm_done
            FROM visit_form vf
            LEFT JOIN engineer_master em ON em.employee_code = vf.employee_code
            WHERE vf.visit_date BETWEEN :s AND :e {sc}
            GROUP BY vf.employee_code, em.employee_name, vf.technician_name, em.service_state
            ORDER BY visits DESC LIMIT 15
        """), params).fetchall()

    tv = int(totals.total_visits or 0)
    solved = int(totals.solved or 0)
    pm = int(totals.pm_done or 0)
    return {
        "date_range": {"start": str(start_d), "end": str(end_d)},
        "summary": {
            "total_visits": tv,
            "unique_sites": int(totals.unique_sites or 0),
            "unique_engineers": int(totals.unique_engineers or 0),
            "ftfr_pct": round(solved / tv * 100, 1) if tv else 0,
            "pm_done_pct": round(pm / tv * 100, 1) if tv else 0,
        },
        "daily": [{"date": str(r.visit_date), "visits": int(r.visits), "solved": int(r.solved), "pm_done": int(r.pm_done)} for r in daily],
        "status_breakdown": [{"status": r.status, "count": int(r.cnt)} for r in status_breakdown],
        "top_engineers": [{"code": r.employee_code, "name": r.name, "state": r.state, "visits": int(r.visits), "ftfr": round(int(r.solved)/int(r.visits)*100,1) if r.visits else 0, "pm_done": int(r.pm_done)} for r in top_engs],
    }


@app.get("/api/visits/browse")
async def visits_browse(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    state: Optional[str] = None,
    employee_code: Optional[str] = None,
    problem_solved: Optional[str] = None,
    pm_done: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    user=Depends(verify_token),
    db: Engine = Depends(get_db)
):
    state_restrict = state_filter(user) or state
    end_d   = date.fromisoformat(end_date)   if end_date   else date.today()
    start_d = date.fromisoformat(start_date) if start_date else end_d - timedelta(days=29)
    offset  = (page - 1) * page_size

    filters = ["vf.visit_date BETWEEN :s AND :e"]
    params: dict = {"s": start_d, "e": end_d, "lim": page_size, "off": offset}
    if state_restrict:
        filters.append("em.service_state = :sr"); params["sr"] = state_restrict
    if employee_code:
        filters.append("vf.employee_code = :ec"); params["ec"] = employee_code
    if problem_solved:
        filters.append("UPPER(vf.problem_solved) = :ps"); params["ps"] = problem_solved.upper()
    if pm_done:
        filters.append("UPPER(vf.pm_done) = :pd"); params["pd"] = pm_done.upper()

    where = " AND ".join(filters)
    with db.connect() as conn:
        total = conn.execute(text(f"""
            SELECT COUNT(*) FROM visit_form vf
            LEFT JOIN engineer_master em ON em.employee_code = vf.employee_code
            WHERE {where}
        """), params).scalar() or 0

        rows = conn.execute(text(f"""
            SELECT vf.id, vf.ticket_id, vf.cs_id, vf.site_name,
                   COALESCE(em.employee_name, vf.technician_name, vf.employee_code) AS engineer_name,
                   COALESCE(em.service_state, '') AS state,
                   vf.visit_date, vf.ticket_status, vf.site_status,
                   vf.problem_solved, vf.pm_done, vf.actual_problem, vf.action_taken
            FROM visit_form vf
            LEFT JOIN engineer_master em ON em.employee_code = vf.employee_code
            WHERE {where}
            ORDER BY vf.visit_date DESC, vf.id DESC
            LIMIT :lim OFFSET :off
        """), params).fetchall()

    return {
        "total": int(total),
        "page": page,
        "page_size": page_size,
        "data": [dict(r._mapping) for r in rows],
    }


@app.get("/api/masters/service-areas")
async def list_service_areas(user=Depends(verify_token), db: Engine = Depends(get_db)):
    with db.connect() as conn:
        rows = conn.execute(text("SELECT service_area_code, service_area_name FROM service_area_master ORDER BY service_area_code")).fetchall()
    return [{"code": r.service_area_code, "name": r.service_area_name} for r in rows]
